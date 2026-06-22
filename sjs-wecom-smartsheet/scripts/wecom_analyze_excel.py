#!/usr/bin/env python3
"""
分析企业微信导出的 Excel 文件
这是最可靠的方案：用户手动导出 Excel，然后用此脚本分析
"""

import os
import sys
import json
from datetime import datetime
from typing import Dict, List, Optional


def analyze_excel(file_path: str) -> Dict:
    """分析 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安装 openpyxl: pip install openpyxl")
        return {}

    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        return {}

    print(f"\n正在分析: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = {
            "file": file_path,
            "sheets": [],
            "timestamp": datetime.now().isoformat()
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {
                "name": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "headers": [],
                "data": []
            }

            # 读取表头
            for row in ws.iter_rows(max_row=1, values_only=True):
                sheet_data["headers"] = [str(cell) if cell else "" for cell in row]

            # 读取数据（最多100行）
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(101, ws.max_row), values_only=True)):
                row_data = [str(cell) if cell else "" for cell in row]
                sheet_data["data"].append(row_data)

            result["sheets"].append(sheet_data)
            print(f"  ✅ Sheet '{sheet_name}': {ws.max_row} 行, {ws.max_column} 列")

        return result

    except Exception as e:
        print(f"❌ 分析失败: {e}")
        return {}


def generate_report(analysis: Dict) -> str:
    """生成分析报告"""
    if not analysis:
        return "分析失败"

    lines = []
    lines.append("# 企业微信智能表格数据分析报告")
    lines.append(f"\n生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"\n文件：{os.path.basename(analysis.get('file', ''))}")

    for sheet in analysis.get("sheets", []):
        lines.append(f"\n## {sheet['name']}")
        lines.append(f"- 行数：{sheet['rows']}")
        lines.append(f"- 列数：{sheet['columns']}")

        if sheet["headers"]:
            lines.append(f"\n### 表头")
            for i, header in enumerate(sheet["headers"], 1):
                if header:
                    lines.append(f"{i}. {header}")

        if sheet["data"]:
            lines.append(f"\n### 数据样本（前10行）")
            for i, row in enumerate(sheet["data"][:10], 1):
                lines.append(f"\n**第{i}行：**")
                for j, cell in enumerate(row):
                    if j < len(sheet["headers"]) and sheet["headers"][j]:
                        lines.append(f"- {sheet['headers'][j]}：{cell[:100]}")

    return "\n".join(lines)


def main():
    """主函数"""
    print("="*60)
    print("企业微信智能表格 Excel 分析工具")
    print("="*60)

    # 检查命令行参数
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # 查找最近下载的 Excel 文件
        download_dir = os.path.expanduser("~/Downloads")
        if os.path.exists(download_dir):
            excel_files = []
            for f in os.listdir(download_dir):
                if f.endswith((".xlsx", ".xls")):
                    full_path = os.path.join(download_dir, f)
                    excel_files.append((full_path, os.path.getmtime(full_path)))

            # 按修改时间排序
            excel_files.sort(key=lambda x: x[1], reverse=True)

            if excel_files:
                print("\n找到以下 Excel 文件：")
                for i, (f, t) in enumerate(excel_files[:10], 1):
                    mtime = datetime.fromtimestamp(t).strftime('%Y-%m-%d %H:%M')
                    print(f"  {i}. {os.path.basename(f)} ({mtime})")

                print("\n请输入文件编号（或直接输入文件路径）：")
                choice = input().strip()

                if choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(excel_files):
                        file_path = excel_files[idx][0]
                    else:
                        print("❌ 无效的编号")
                        return
                else:
                    file_path = choice
            else:
                print("\n未找到 Excel 文件")
                print("请手动输入文件路径：")
                file_path = input().strip()
        else:
            print("\n请手动输入 Excel 文件路径：")
            file_path = input().strip()

    if not file_path:
        print("❌ 未指定文件")
        return

    # 分析 Excel
    analysis = analyze_excel(file_path)

    if analysis:
        # 生成报告
        report = generate_report(analysis)

        # 使用 workspace 目录存放输出
        workspace_dir = os.path.expanduser(
            "~/work_space/ai/skill/sjs-skills/sjs-wecom-smartsheet/workspace"
        )
        os.makedirs(workspace_dir, exist_ok=True)

        # 按日期创建子目录
        from datetime import datetime
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(workspace_dir, date_str)
        os.makedirs(output_dir, exist_ok=True)

        report_file = os.path.join(
            output_dir,
            f"analysis_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
        )
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report)

        # 保存原始数据
        data_file = os.path.join(output_dir, "excel_data.json")
        with open(data_file, 'w', encoding='utf-8') as f:
            json.dump(analysis, f, ensure_ascii=False, indent=2)

        print("\n" + "="*60)
        print("✅ 分析完成！")
        print("="*60)
        print(f"\n输出文件：")
        print(f"  - 报告: {report_file}")
        print(f"  - 数据: {data_file}")

        # 显示报告预览
        print("\n" + "="*60)
        print("📊 报告预览")
        print("="*60)
        print(report[:2000])  # 显示前2000字符

    else:
        print("\n❌ 分析失败")


if __name__ == "__main__":
    main()
